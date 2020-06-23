import requests
from bs4 import BeautifulSoup
import smtplib
import openpyxl
from openpyxl import Workbook
from openpyxl import styles
from datetime import date
from win10toast import ToastNotifier
from amd import *
from intel import *
from nvidea import *
excel = Workbook()

#CPU
cpu = excel.create_sheet('CPU', 0)
cpu.title = 'CPU'


cpu['A1'] = 'AMD:'
cpu['A2'] = 'CPU:'
cpu['B1'] = 'Nome:'
cpu['B2'] = 'R5 1600AF'
cpu['D1'] = 'Kabum:'
cpu['C1'] = 'Data:'
cpu['E1'] = 'Status:'
cpu['C2'] = '{}'.format(date.today())
cpu['D2'] = 'R${}'.format(price_kabum_R51600AF)
cpu['D2'].hyperlink = '{}'.format(kabum_R51600AF)
cpu['E2'] = '{}'.format(status_kabum_R51600AF)
cpu['A3'] = 'CPU:'
cpu['B3'] = 'R3 2200G'
cpu['D3'] = '{}'.format(price_kabum_R32200G)
cpu['C3'] = '{}'.format(date.today())
cpu['E3'] = '{}'.format(status_kabum_R32200G)
cpu['A4'] = 'CPU:'
cpu['B4'] = 'R5 2600'
cpu['D4'] = 'R${}'.format(price_kabum_R52600)
cpu['C4'] = '{}'.format(date.today())
cpu['E4'] = '{}'.format(status_kabum_R52600)
cpu['A5'] = 'CPU:'
cpu['B5'] = 'R7 2700'
cpu['D5'] = 'R${}'.format(price_kabum_R72700)
cpu['C5'] = '{}'.format(date.today())
cpu['E5'] = '{}'.format(status_kabum_R72700)
cpu['A6'] = 'CPU:'
cpu['B6'] = 'R5 3600'
cpu['D6'] = 'R${}'.format(price_kabum_R53600)
cpu['C6'] = '{}'.format(date.today())
cpu['E6'] = '{}'.format(status_kabum_R53600)
cpu['A7'] = 'CPU:'
cpu['B7'] = 'R7 3700X'
cpu['D7'] = 'R${}'.format(price_kabum_R73700X)
cpu['C7'] = '{}'.format(date.today())
cpu['E7'] = '{}'.format(status_kabum_R73700X)
cpu['A8'] = 'CPU:'
cpu['B8'] = 'R7 3800X'
cpu['D8'] = 'R${}'.format(price_kabum_R73800X)
cpu['C8'] = '{}'.format(date.today())
cpu['E8'] = '{}'.format(status_kabum_R73800X)
cpu['A9'] = 'CPU:'
cpu['B9'] = 'R9 3900X'
cpu['D9'] = 'R${}'.format(price_kabum_R93900X)
cpu['C9'] = '{}'.format(date.today())
cpu['E9'] = '{}'.format(status_kabum_R93900X)
cpu['A10'] = 'CPU:'
cpu['B10'] = 'R3 3300X'
cpu['D10'] = 'R${}'.format(price_kabum_R33300X)
cpu['C10'] = '{}'.format(date.today())
cpu['E10'] = '{}'.format(status_kabum_R33300X)
cpu['A11'] = 'CPU:'
cpu['B11'] = 'R3 3200G'
cpu['D11'] = 'R${}'.format(price_kabum_R33200G)
cpu['C11'] = '{}'.format(date.today())
cpu['E11'] = '{}'.format(status_kabum_R33200G)
cpu['A12'] = 'CPU:'
cpu['B12'] = 'R5 3600X'
cpu['D12'] = 'R${}'.format(price_kabum_R53600X)
cpu['C12'] = '{}'.format(date.today())
cpu['E12'] = '{}'.format(status_kabum_R53600X)


cpu['A18'] = 'Intel:'
cpu['B18'] = 'Nome:'
cpu['C18'] = 'Data:'
cpu['D18'] = 'Kabum:'
cpu['A18'] = 'CPU:'
cpu['E18'] = 'Status:'
cpu['B19'] = 'i3 9100F'
cpu['D19'] = 'R${}'.format(price_kabum_i39100F)
cpu['C19'] = '{}'.format(date.today())
cpu['E19'] = '{}'.format(status_kabum_i39100F)
cpu['A20'] = 'CPU:'
cpu['B20'] = 'i5 9400F'
cpu['D20'] = 'R${}'.format(price_kabum_i59400F)
cpu['C20'] = '{}'.format(date.today())
cpu['E20'] = '{}'.format(status_kabum_i59400F)
cpu['A21'] = 'CPU:'
cpu['B21'] = 'i5 9600K'
cpu['D21'] = 'R${}'.format(price_kabum_i59600K)
cpu['C21'] = '{}'.format(date.today())
cpu['C21'] = '{}'.format(status_kabum_i59600K)
cpu['A22'] = 'CPU:'
cpu['B22'] = 'i5 9600KF'
cpu['D22'] = 'R${}'.format(price_kabum_i59600KF)
cpu['C22'] = '{}'.format(date.today())
cpu['E22'] = '{}'.format(status_kabum_i59600KF)
cpu['A23'] = 'CPU:'
cpu['B23'] = 'i7 9700K'
cpu['D23'] = 'R${}'.format(price_kabum_i79700K)
cpu['C23'] = '{}'.format(date.today())
cpu['E24'] = '{}'.format(status_kabum_i79700K)
cpu['A24'] = 'CPU:'
cpu['B24'] = 'i9 9900K'
cpu['D24'] = 'R${}'.format(price_kabum_i99900K)
cpu['C24'] = '{}'.format(date.today())
cpu['E24'] = '{}'.format(status_kabum_i99900K)
cpu['A25'] = 'CPU:'
cpu['B25'] = 'i3 10100'
cpu['D25'] = 'R${}'.format(price_kabum_i310100)
cpu['C25'] = '{}'.format(date.today())
cpu['E25'] = '{}'.format(status_kabum_i310100)
cpu['A26'] = 'CPU:'
cpu['B26'] = 'i5 10400'
cpu['D26'] = 'R${}'.format(price_kabum_i510400)
cpu['C26'] = '{}'.format(date.today())
cpu['E26'] = '{}'.format(status_kabum_i510400)
cpu['A27'] = 'CPU:'
cpu['B27'] = 'i5 10400F'
cpu['D27'] = 'R${}'.format(price_kabum_i510400F)
cpu['C27'] = '{}'.format(date.today())
cpu['E27'] = '{}'.format(status_kabum_i510400F)
cpu['A28'] = 'CPU:'
cpu['B28'] = 'i7 10700'
cpu['D28'] = 'R${}'.format(price_kabum_i710700)
cpu['C28'] = '{}'.format(date.today())
cpu['E28'] = '{}'.format(status_kabum_i710700)
cpu['A29'] = 'CPU:'
cpu['B29'] = 'i7 10700K'
cpu['D29'] = 'R${}'.format(price_kabum_i710700K)
cpu['C29'] = '{}'.format(date.today())
cpu['E29'] = '{}'.format(status_kabum_i710700K)
cpu['A30'] = 'CPU:'
cpu['B30'] = 'i9 10900K'
cpu['D30'] = 'R${}'.format(price_kabum_i910900K)
cpu['C30'] = '{}'.format(date.today())
cpu['E30'] = '{}'.format(status_kabum_i910900K)



#GPU

gpu = excel.create_sheet('GPU', 1)
gpu.title = 'GPU'


gpu['A1'] = 'AMD:'
gpu['A2'] = 'GPU:'
gpu['B1'] = 'Nome:'
gpu['B2'] = 'RX570 4GB'
gpu['D1'] = 'Kabum:'
gpu['C1'] = 'Data:'
gpu['E1'] = 'Modelo:'
gpu['G1'] = 'Status:'
gpu['C2'] = '{}'.format(date.today())
gpu['D2'] = 'R${}'.format(price_kabum_Rx5704gb_PhantomGaming)
gpu['E2'] = 'Phantom Gaming D'
gpu['F1'] = 'Fabricante:'
gpu['F2'] = 'Asrock'
gpu['G2'] = '{}'.format(status_kabum_Rx5704gb_PhantomGaming)
gpu['A3'] = 'GPU:'
gpu['B3'] = 'RX570 8GB'
gpu['C3'] = '{}'.format(date.today())
gpu['D3'] = 'R${}'.format(price_kabum_Rx5708gb_PhantomGaming)
gpu['E3'] = 'Phantom Gaming D'
gpu['F3'] = 'Asrock'
gpu['G3'] = '{}'.format(status_kabum_Rx5708gb_PhantomGaming)
gpu['A4'] = 'GPU:'
gpu['B4'] = 'RX570 4gb'
gpu['C4'] = '{}'.format(date.today())
gpu['D4'] = 'R${}'.format(price_kabum_Rx5704gb_Gaming4G)
gpu['E4'] = 'Gaming 4G'
gpu['F4'] = 'Gigabyte'
gpu['A5'] = 'GPU:'
gpu['B5'] = 'RX570 8gb'
gpu['C5'] = '{}'.format(date.today())
gpu['D5'] = 'R${}'.format(price_kabum_Rx5708gb_Gaming8G)
gpu['E5'] = 'Gaming 8G'
gpu['F5'] = 'Gigabyte'
gpu['G5'] = '{}'.format(status_kabum_Rx5708gb_Gaming8G)
gpu['A6'] = 'GPU:'
gpu['B6'] = 'RX570 4gb'
gpu['C6'] = '{}'.format(date.today())
gpu['D6'] = 'R${}'.format(price_kabum_Rx5704gb_XFXRSXXX)
gpu['E6'] = 'RS XXX OC'
gpu['F6'] = 'XFX'
gpu['G6'] = '{}'.format(status_kabum_Rx5704gb_XFXRSXXX)
gpu['A7'] = 'GPU:'
gpu['B7'] = 'RX590'
gpu['C7'] = '{}'.format(date.today())
gpu['D7'] = 'R${}'.format(price_kabum_Rx590_Gaming8G)
gpu['E7'] = 'Gaming8G'
gpu['F7'] = 'Gigabyte'
gpu['G7'] = '{}'.format(status_kabum_Rx590_Gaming8G)
gpu['A8'] = 'GPU:'
gpu['B8'] = 'RX5500XT 4gb'
gpu['C8'] = '{}'.format(date.today())
gpu['D8'] = 'R${}'.format(price_kabum_Rx5500XT4gb_GamingOC)
gpu['E8'] = 'Gaming OC'
gpu['F8'] = 'Gigabyte'
gpu['G8'] = '{}'.format(status_kabum_Rx5500XT4gb_GamingOC)
gpu['A9'] = 'GPU:'
gpu['B9'] = 'RX5500XT 8gb'
gpu['C9'] = '{}'.format(date.today())
gpu['D9'] = 'R${}'.format(price_kabum_Rx5500XT8gb_Pulse)
gpu['E9'] = 'Pulse'
gpu['F9'] = 'Sapphire'
gpu['G9'] = '{}'.format(status_kabum_Rx5500XT8gb_Pulse)






gpu['A19'] = 'Nvidea:'
gpu['B19'] = 'Nome:'
gpu['C19'] = 'Data:'
gpu['D19'] = 'Kabum:'
gpu['E19'] = 'Modelo:'
gpu['F19'] = 'Fabricante:'
gpu['G19'] = 'Status:'
gpu['A20'] = 'GPU:'
gpu['B20'] = 'GTX 1650 Super'
gpu['C20'] = '{}'.format(date.today())
gpu['D20'] = 'R${}'.format(price_kabum_GTX1650Super_XCBlackGaming)
gpu['E20'] = 'XC Black Gaming'
gpu['F20'] = 'EVGA'
gpu['G20'] = '{}'.format(status_kabum_GTX1650Super_XCBlackGaming)
gpu['A21'] = 'GPU:'
gpu['B21'] = 'GTX 1660'
gpu['C21'] = '{}'.format(date.today())
gpu['D21'] = 'R${}'.format(price_kabum_GTX1660_GigabyteOC)
gpu['E21'] = 'Gigabyte'
gpu['F21'] = 'Gigabyte'
gpu['G21'] = '{}'.format(status_kabum_GTX1660_GigabyteOC)
gpu['A22'] = 'GPU:'
gpu['B22'] = 'GTX 1660 Super'
gpu['C22'] = '{}'.format(date.today())
gpu['D22'] = 'R${}'.format(price_kabum_GTX1660Super_GigabyteOC)
gpu['E22'] = 'Gigabyte'
gpu['F22'] = 'Gigabyte'
gpu['G22'] = '{}'.format(status_kabum_GTX1660Super_GigabyteOC)
gpu['A23'] = 'GPU:'
gpu['B23'] = 'GTX 1660 Ti'
gpu['C23'] = '{}'.format(date.today())
gpu['D23'] = 'R${}'.format(price_kabum_GTX1660Ti_GamingX)
gpu['E23'] = 'Gaming X'
gpu['F23'] = 'MSI'
gpu['G23'] = '{}'.format(status_kabum_GTX1660Ti_GamingX)
gpu['A24'] = 'GPU:'
gpu['B24'] = 'RTX 2060'
gpu['C24'] = '{}'.format(date.today())
gpu['D24'] = 'R${}'.format(price_kabum_RTX2060_Gaming)
gpu['E24'] = 'Gaming'
gpu['F24'] = 'EVGA'
gpu['G24'] = '{}'.format(status_kabum_RTX2060_Gaming)
gpu['A25'] = 'GPU:'
gpu['B25'] = 'RTX 2060'
gpu['C25'] = '{}'.format(date.today())
gpu['D25'] = 'R${}'.format(price_kabum_RTX2060_GamingOCPro)
gpu['E25'] = 'Gaming OC Pro'
gpu['F25'] = 'Gigabyte'
gpu['G25'] = '{}'.format(status_kabum_RTX2060_GamingOCPro)
gpu['A26'] = 'GPU:'
gpu['B26'] = 'RTX 2060 Super'
gpu['C26'] = '{}'.format(date.today())
gpu['D26'] = 'R${}'.format(price_kabum_RTX2060Super_ROGStrix)
gpu['E26'] = 'ROG Strix'
gpu['F26'] = 'Asus'
gpu['G26'] = '{}'.format(status_kabum_RTX2060Super_ROGStrix)

excel.save('{}.xlsx'.format(date.today()))





























