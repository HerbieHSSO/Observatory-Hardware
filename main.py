import requests
from bs4 import BeautifulSoup
import smtplib
import openpyxl
from threading import Thread
from multiprocessing import Process
from time import sleep
import time
from openpyxl import Workbook
from openpyxl import styles
from datetime import date
from win10toast import ToastNotifier
import tweepy
from amd import *
from intel import *
from nvidea import *






auth = tweepy.OAuthHandler('') # your twitter auth
auth.set_access_token('')# your twitter token


api = tweepy.API(auth)




try:
    api.verify_credentials()
    print("Autenticação aceita")
except:
    print("Um erro ocorreu durante a autenticação.")





excel = Workbook()

#CPU
cpu = excel.create_sheet('CPU', 0)
cpu.title = 'CPU'


cpu['A1'] = 'AMD:'
cpu['A2'] = 'CPU:'
cpu['B1'] = 'Nome:'
cpu['B2'] = 'R5 1600AF'
cpu['D1'] = 'Kabum:'
cpu['C1'] = 'Data:'
cpu['E1'] = 'Status:'
cpu['C2'] = '{}'.format(date.today())
cpu['D2'] = 'R${}'.format(price_kabum_R51600AF)
cpu['D2'].hyperlink = '{}'.format(kabum_R51600AF)
cpu['E2'] = '{}'.format(status_kabum_R51600AF)
cpu['A3'] = 'CPU:'
cpu['B3'] = 'R3 2200G'
cpu['D3'] = '{}'.format(price_kabum_R32200G)
cpu['C3'] = '{}'.format(date.today())
cpu['E3'] = '{}'.format(status_kabum_R32200G)
cpu['A4'] = 'CPU:'
cpu['B4'] = 'R5 2600'
cpu['D4'] = 'R${}'.format(price_kabum_R52600)
cpu['C4'] = '{}'.format(date.today())
cpu['E4'] = '{}'.format(status_kabum_R52600)
cpu['A5'] = 'CPU:'
cpu['B5'] = 'R7 2700'
cpu['D5'] = 'R${}'.format(price_kabum_R72700)
cpu['C5'] = '{}'.format(date.today())
cpu['E5'] = '{}'.format(status_kabum_R72700)
cpu['A6'] = 'CPU:'
cpu['B6'] = 'R5 3600'
cpu['D6'] = 'R${}'.format(price_kabum_R53600)
cpu['C6'] = '{}'.format(date.today())
cpu['E6'] = '{}'.format(status_kabum_R53600)
cpu['A7'] = 'CPU:'
cpu['B7'] = 'R7 3700X'
cpu['D7'] = 'R${}'.format(price_kabum_R73700X)
cpu['C7'] = '{}'.format(date.today())
cpu['E7'] = '{}'.format(status_kabum_R73700X)
cpu['A8'] = 'CPU:'
cpu['B8'] = 'R7 3800X'
cpu['D8'] = 'R${}'.format(price_kabum_R73800X)
cpu['C8'] = '{}'.format(date.today())
cpu['E8'] = '{}'.format(status_kabum_R73800X)
cpu['A9'] = 'CPU:'
cpu['B9'] = 'R9 3900X'
cpu['D9'] = 'R${}'.format(price_kabum_R93900X)
cpu['C9'] = '{}'.format(date.today())
cpu['E9'] = '{}'.format(status_kabum_R93900X)
cpu['A10'] = 'CPU:'
cpu['B10'] = 'R3 3300X'
cpu['D10'] = 'R${}'.format(price_kabum_R33300X)
cpu['C10'] = '{}'.format(date.today())
cpu['E10'] = '{}'.format(status_kabum_R33300X)
cpu['A11'] = 'CPU:'
cpu['B11'] = 'R3 3200G'
cpu['D11'] = 'R${}'.format(price_kabum_R33200G)
cpu['C11'] = '{}'.format(date.today())
cpu['E11'] = '{}'.format(status_kabum_R33200G)
cpu['A12'] = 'CPU:'
cpu['B12'] = 'R5 3600X'
cpu['D12'] = 'R${}'.format(price_kabum_R53600X)
cpu['C12'] = '{}'.format(date.today())
cpu['E12'] = '{}'.format(status_kabum_R53600X)


cpu['A18'] = 'Intel:'
cpu['B18'] = 'Nome:'
cpu['C18'] = 'Data:'
cpu['D18'] = 'Kabum:'
cpu['A18'] = 'CPU:'
cpu['E18'] = 'Status:'
cpu['B19'] = 'i3 9100F'
cpu['D19'] = 'R${}'.format(price_kabum_i39100F)
cpu['C19'] = '{}'.format(date.today())
cpu['E19'] = '{}'.format(status_kabum_i39100F)
cpu['A20'] = 'CPU:'
cpu['B20'] = 'i5 9400F'
cpu['D20'] = 'R${}'.format(price_kabum_i59400F)
cpu['C20'] = '{}'.format(date.today())
cpu['E20'] = '{}'.format(status_kabum_i59400F)
cpu['A21'] = 'CPU:'
cpu['B21'] = 'i5 9600K'
cpu['D21'] = 'R${}'.format(price_kabum_i59600K)
cpu['C21'] = '{}'.format(date.today())
cpu['C21'] = '{}'.format(status_kabum_i59600K)
cpu['A22'] = 'CPU:'
cpu['B22'] = 'i5 9600KF'
cpu['D22'] = 'R${}'.format(price_kabum_i59600KF)
cpu['C22'] = '{}'.format(date.today())
cpu['E22'] = '{}'.format(status_kabum_i59600KF)
cpu['A23'] = 'CPU:'
cpu['B23'] = 'i7 9700K'
cpu['D23'] = 'R${}'.format(price_kabum_i79700K)
cpu['C23'] = '{}'.format(date.today())
cpu['E24'] = '{}'.format(status_kabum_i79700K)
cpu['A24'] = 'CPU:'
cpu['B24'] = 'i9 9900K'
cpu['D24'] = 'R${}'.format(price_kabum_i99900K)
cpu['C24'] = '{}'.format(date.today())
cpu['E24'] = '{}'.format(status_kabum_i99900K)
cpu['A25'] = 'CPU:'
cpu['B25'] = 'i3 10100'
cpu['D25'] = 'R${}'.format(price_kabum_i310100)
cpu['C25'] = '{}'.format(date.today())
cpu['E25'] = '{}'.format(status_kabum_i310100)
cpu['A26'] = 'CPU:'
cpu['B26'] = 'i5 10400'
cpu['D26'] = 'R${}'.format(price_kabum_i510400)
cpu['C26'] = '{}'.format(date.today())
cpu['E26'] = '{}'.format(status_kabum_i510400)
cpu['A27'] = 'CPU:'
cpu['B27'] = 'i5 10400F'
cpu['D27'] = 'R${}'.format(price_kabum_i510400F)
cpu['C27'] = '{}'.format(date.today())
cpu['E27'] = '{}'.format(status_kabum_i510400F)
cpu['A28'] = 'CPU:'
cpu['B28'] = 'i7 10700'
cpu['D28'] = 'R${}'.format(price_kabum_i710700)
cpu['C28'] = '{}'.format(date.today())
cpu['E28'] = '{}'.format(status_kabum_i710700)
cpu['A29'] = 'CPU:'
cpu['B29'] = 'i7 10700K'
cpu['D29'] = 'R${}'.format(price_kabum_i710700K)
cpu['C29'] = '{}'.format(date.today())
cpu['E29'] = '{}'.format(status_kabum_i710700K)
cpu['A30'] = 'CPU:'
cpu['B30'] = 'i9 10900K'
cpu['D30'] = 'R${}'.format(price_kabum_i910900K)
cpu['C30'] = '{}'.format(date.today())
cpu['E30'] = '{}'.format(status_kabum_i910900K)



#GPU

gpu = excel.create_sheet('GPU', 1)
gpu.title = 'GPU'


gpu['A1'] = 'AMD:'
gpu['A2'] = 'GPU:'
gpu['B1'] = 'Nome:'
gpu['B2'] = 'RX570 4GB'
gpu['D1'] = 'Kabum:'
gpu['C1'] = 'Data:'
gpu['E1'] = 'Modelo:'
gpu['G1'] = 'Status:'
gpu['C2'] = '{}'.format(date.today())
gpu['D2'] = 'R${}'.format(price_kabum_Rx5704gb_PhantomGaming)
gpu['E2'] = 'Phantom Gaming D'
gpu['F1'] = 'Fabricante:'
gpu['F2'] = 'Asrock'
gpu['G2'] = '{}'.format(status_kabum_Rx5704gb_PhantomGaming)
gpu['A3'] = 'GPU:'
gpu['B3'] = 'RX570 8GB'
gpu['C3'] = '{}'.format(date.today())
gpu['D3'] = 'R${}'.format(price_kabum_Rx5708gb_PhantomGaming)
gpu['E3'] = 'Phantom Gaming D'
gpu['F3'] = 'Asrock'
gpu['G3'] = '{}'.format(status_kabum_Rx5708gb_PhantomGaming)
gpu['A4'] = 'GPU:'
gpu['B4'] = 'RX570 4gb'
gpu['C4'] = '{}'.format(date.today())
gpu['D4'] = 'R${}'.format(price_kabum_Rx5704gb_Gaming4G)
gpu['E4'] = 'Gaming 4G'
gpu['F4'] = 'Gigabyte'
gpu['A5'] = 'GPU:'
gpu['B5'] = 'RX570 8gb'
gpu['C5'] = '{}'.format(date.today())
gpu['D5'] = 'R${}'.format(price_kabum_Rx5708gb_Gaming8G)
gpu['E5'] = 'Gaming 8G'
gpu['F5'] = 'Gigabyte'
gpu['G5'] = '{}'.format(status_kabum_Rx5708gb_Gaming8G)
gpu['A6'] = 'GPU:'
gpu['B6'] = 'RX570 4gb'
gpu['C6'] = '{}'.format(date.today())
gpu['D6'] = 'R${}'.format(price_kabum_Rx5704gb_XFXRSXXX)
gpu['E6'] = 'RS XXX OC'
gpu['F6'] = 'XFX'
gpu['G6'] = '{}'.format(status_kabum_Rx5704gb_XFXRSXXX)
gpu['A7'] = 'GPU:'
gpu['B7'] = 'RX590'
gpu['C7'] = '{}'.format(date.today())
gpu['D7'] = 'R${}'.format(price_kabum_Rx590_Gaming8G)
gpu['E7'] = 'Gaming8G'
gpu['F7'] = 'Gigabyte'
gpu['G7'] = '{}'.format(status_kabum_Rx590_Gaming8G)
gpu['A8'] = 'GPU:'
gpu['B8'] = 'RX5500XT 4gb'
gpu['C8'] = '{}'.format(date.today())
gpu['D8'] = 'R${}'.format(price_kabum_Rx5500XT4gb_GamingOC)
gpu['E8'] = 'Gaming OC'
gpu['F8'] = 'Gigabyte'
gpu['G8'] = '{}'.format(status_kabum_Rx5500XT4gb_GamingOC)
gpu['A9'] = 'GPU:'
gpu['B9'] = 'RX5500XT 8gb'
gpu['C9'] = '{}'.format(date.today())
gpu['D9'] = 'R${}'.format(price_kabum_Rx5500XT8gb_Pulse)
gpu['E9'] = 'Pulse'
gpu['F9'] = 'Sapphire'
gpu['G9'] = '{}'.format(status_kabum_Rx5500XT8gb_Pulse)






gpu['A19'] = 'Nvidea:'
gpu['B19'] = 'Nome:'
gpu['C19'] = 'Data:'
gpu['D19'] = 'Kabum:'
gpu['E19'] = 'Modelo:'
gpu['F19'] = 'Fabricante:'
gpu['G19'] = 'Status:'
gpu['A20'] = 'GPU:'
gpu['B20'] = 'GTX 1650 Super'
gpu['C20'] = '{}'.format(date.today())
gpu['D20'] = 'R${}'.format(price_kabum_GTX1650Super_XCBlackGaming)
gpu['E20'] = 'XC Black Gaming'
gpu['F20'] = 'EVGA'
gpu['G20'] = '{}'.format(status_kabum_GTX1650Super_XCBlackGaming)
gpu['A21'] = 'GPU:'
gpu['B21'] = 'GTX 1660'
gpu['C21'] = '{}'.format(date.today())
gpu['D21'] = 'R${}'.format(price_kabum_GTX1660_GigabyteOC)
gpu['E21'] = 'Gigabyte'
gpu['F21'] = 'Gigabyte'
gpu['G21'] = '{}'.format(status_kabum_GTX1660_GigabyteOC)
gpu['A22'] = 'GPU:'
gpu['B22'] = 'GTX 1660 Super'
gpu['C22'] = '{}'.format(date.today())
gpu['D22'] = 'R${}'.format(price_kabum_GTX1660Super_GigabyteOC)
gpu['E22'] = 'Gigabyte'
gpu['F22'] = 'Gigabyte'
gpu['G22'] = '{}'.format(status_kabum_GTX1660Super_GigabyteOC)
gpu['A23'] = 'GPU:'
gpu['B23'] = 'GTX 1660 Ti'
gpu['C23'] = '{}'.format(date.today())
gpu['D23'] = 'R${}'.format(price_kabum_GTX1660Ti_GamingX)
gpu['E23'] = 'Gaming X'
gpu['F23'] = 'MSI'
gpu['G23'] = '{}'.format(status_kabum_GTX1660Ti_GamingX)
gpu['A24'] = 'GPU:'
gpu['B24'] = 'RTX 2060'
gpu['C24'] = '{}'.format(date.today())
gpu['D24'] = 'R${}'.format(price_kabum_RTX2060_Gaming)
gpu['E24'] = 'Gaming'
gpu['F24'] = 'EVGA'
gpu['G24'] = '{}'.format(status_kabum_RTX2060_Gaming)
gpu['A25'] = 'GPU:'
gpu['B25'] = 'RTX 2060'
gpu['C25'] = '{}'.format(date.today())
gpu['D25'] = 'R${}'.format(price_kabum_RTX2060_GamingOCPro)
gpu['E25'] = 'Gaming OC Pro'
gpu['F25'] = 'Gigabyte'
gpu['G25'] = '{}'.format(status_kabum_RTX2060_GamingOCPro)
gpu['A26'] = 'GPU:'
gpu['B26'] = 'RTX 2060 Super'
gpu['C26'] = '{}'.format(date.today())
gpu['D26'] = 'R${}'.format(price_kabum_RTX2060Super_ROGStrix)
gpu['E26'] = 'ROG Strix'
gpu['F26'] = 'Asus'
gpu['G26'] = '{}'.format(status_kabum_RTX2060Super_ROGStrix)

excel.save('Kabum{}.xlsx'.format(date.today()))
def R51600AF():
    time.sleep(5)
    while True:
        if status_kabum_R51600AF != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R51600AF))
        else:

            kabum_R51600AF_Promocao = api.update_status('''

PROMOÇÃO:

R5 1600(AF)

''' + oldPrice_kabum_R51600AF + '''
''' + newPrice_kabum_R51600AF + '''


Preço: R$''' + price_kabum_R51600AF + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R51600AF + '''
                                            ''')
            break
            while True:
          
                if status_kabum_R51600AF == soup_kabum_R51600AF.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
            
                    api.destroy_status(kabum_R51600AF_Promocao)

    


def R32200G():
    time.sleep(5)
    while True:
        if status_kabum_R32200G != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R32200G))


        else:
            kabum_R32200G_Promocao = api.update_status('''

PROMOÇÃO:

R3 2200G

''' + oldPrice_kabum_R32200G + '''
''' + newPrice_kabum_R32200G + '''


Preço: R$''' + price_kabum_R32200G + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R32200G + '''

                                                                    ''')
            break
            while True:
          
                if status_kabum_R32200G == soup_kabum_R32200G.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R32200G_Promocao)




def R52600():
    
    
    while True:
        if status_kabum_R52600 != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R52600))


        else:
            kabum_R52600_Promocao = api.update_status('''

PROMOÇÃO:

R5 2600

''' + oldPrice_kabum_R52600 + '''
''' + newPrice_kabum_R52600 + '''


Preço: R$''' + price_kabum_R52600 + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R52600 + '''



                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R52600 == soup_kabum_R52600.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R52600_Promocao)


    
def R53600():
    while True:
        if status_kabum_R53600 != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R53600))


        else:
            kabum_R53600_Promocao = api.update_status('''

PROMOÇÃO:

R5 3600

''' + oldPrice_kabum_R53600 + '''
''' + newPrice_kabum_R53600 + '''


Preço: R$''' + price_kabum_R53600 + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R53600 + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R53600 == soup_kabum_R53600.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R53600_Promocao)
                    
def R72700():
    while True:
        if status_kabum_R72700 != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R72700))


        else:
            kabum_R72700_Promocao = api.update_status('''

PROMOÇÃO:

R7 2700

''' + oldPrice_kabum_R72700 + '''
''' + newPrice_kabum_R72700 + '''


Preço: R$''' + price_kabum_R72700 + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R72700 + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R72700 == soup_kabum_R72700.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R72700_Promocao)

def R73700X():
    while True:
        if status_kabum_R73700X != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R73700X))


        else:
            kabum_R73700X_Promocao = api.update_status('''

PROMOÇÃO:

R7 3700X

''' + oldPrice_kabum_R73700X + '''
''' + newPrice_kabum_R73700X + '''


Preço: R$''' + price_kabum_R73700X + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R73700X + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R73700X == soup_kabum_R73700X.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R73700X_Promocao)


def R73800X():
    while True:
        if status_kabum_R73800X != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R73800X))


        else:
            kabum_R73800X_Promocao = api.update_status('''

PROMOÇÃO:

R7 3800X

''' + oldPrice_kabum_R73800X + '''
''' + newPrice_kabum_R73800X + '''


Preço: R$''' + price_kabum_R73800X + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R73800X + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R73800X == soup_kabum_R73800X.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R73800X_Promocao)

def R93900X():
    while True:
        if status_kabum_R93900X != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R93900X))


        else:
            kabum_R93900X_Promocao = api.update_status('''

PROMOÇÃO:

R9 3900X

''' + oldPrice_kabum_R93900X + '''
''' + newPrice_kabum_R93900X + '''


Preço: R$''' + price_kabum_R93900X + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R93900X + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R93900X == soup_kabum_R93900X.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R93900X_Promocao)


def R33300X():
    while True:
        if status_kabum_R33300X != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R33300X))


        else:
            kabum_R33300X_Promocao = api.update_status('''

PROMOÇÃO:

R3 3300X

''' + oldPrice_kabum_R33300X + '''
''' + newPrice_kabum_R33300X + '''


Preço: R$''' + price_kabum_R33300X + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R33300X + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R33300X == soup_kabum_R33300X.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R33300X_Promocao)
def R33200G():
    while True:
        if status_kabum_R33200G != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R33200G))


        else:
            kabum_R33200G_Promocao = api.update_status('''

PROMOÇÃO:

R3 3200G

''' + oldPrice_kabum_R33200G + '''
''' + newPrice_kabum_R33200G + '''


Preço: R$''' + price_kabum_R33200G + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R33200G + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R33200G == soup_kabum_R33200G.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R33200G_Promocao)

def R53600X():
    while True:
        if status_kabum_R53600X != 'PROMOÇÃO':
            print('{}'.format(status_kabum_R53600X))


        else:
            kabum_R53600X_Promocao = api.update_status('''

PROMOÇÃO:

R5 3600X

''' + oldPrice_kabum_R53600X + '''
''' + newPrice_kabum_R53600X + '''


Preço: R$''' + price_kabum_R53600X + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_R53600X + '''

                                                                    ''')
            
            break
            while True:
          
                if status_kabum_R53600X == soup_kabum_R53600X.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_R53600X_Promocao)


def Rx5704gb_PhantomGaming():
    while True:
        if status_kabum_Rx5704gb_PhantomGaming != 'PROMOÇÃO':
            print('{}'.format(status_kabum_Rx5704gb_PhantomGaming))


        else:
            kabum_Rx5704gb_PhantomGaming_Promocao = api.update_status('''
PROMOÇÃO:

Rx570 4gb (PhantomGaming - Asrock)

''' + oldPrice_kabum_Rx5704gb_PhantomGaming + '''
''' + newPrice_kabum_Rx5704gb_PhantomGaming + '''


Preço: R$''' + price_kabum_Rx5704gb_PhantomGaming + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx5704gb_PhantomGaming + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx5704gb_PhantomGaming == soup_kabum_Rx5704gb_PhantomGaming.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx5704gb_PhantomGaming_Promocao)    
    

def Rx5708gb_PhantomGaming():
    while True:
        if status_kabum_Rx5708gb_PhantomGaming != 'PROMOÇÃO':
            print('{}'.format(status_kabum_Rx5708gb_PhantomGaming))


        else:
            kabum_Rx5708gb_PhantomGaming_Promocao = api.update_status('''

PROMOÇÃO:

Rx570 8gb (PhantomGaming - Asrock)

''' + oldPrice_kabum_Rx5708gb_PhantomGaming + '''
''' + newPrice_kabum_Rx5708gb_PhantomGaming + '''


Preço: R$''' + price_kabum_Rx5708gb_PhantomGaming + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx5708gb_PhantomGaming + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx5708gb_PhantomGaming == soup_kabum_Rx5708gb_PhantomGaming.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx5708gb_PhantomGaming_Promocao)    
    

def Rx5704gb_Gaming4G():
    while True:
        if status_kabum_Rx5704gb_Gaming4G != 'PROMOÇÃO':
            print('{}'.format(status_kabum_Rx5704gb_Gaming4G))


        else:
            kabum_Rx5704gb_Gaming4G_Promocao = api.update_status('''

PROMOÇÃO:

Rx570 4gb (Gaming4G - Gigabyte)

''' + oldPrice_kabum_Rx5704gb_Gaming4G + '''
''' + newPrice_kabum_Rx5704gb_Gaming4G + '''


Preço: R$''' + price_kabum_Rx5704gb_Gaming4G + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx5704gb_Gaming4G + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx5704gb_Gaming4G == soup_kabum_Rx5704gb_Gaming4G.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx5704gb_Gaming4G_Promocao)    
    
def Rx5708gb_Gaming8G():
    while True:
        if status_kabum_Rx5708gb_Gaming8G != 'PROMOÇÃO':
            print('{}'.format(status_kabum_Rx5708gb_Gaming8G))


        else:
            kabum_Rx5708gb_Gaming8G_Promocao = api.update_status('''

PROMOÇÃO:

Rx570 8gb (Gaming8G - Gigabyte)

''' + oldPrice_kabum_Rx5708gb_Gaming8G + '''
''' + newPrice_kabum_Rx5708gb_Gaming8G + '''


Preço: R$''' + price_kabum_Rx5708gb_Gaming8G + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx5708gb_Gaming8G + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx5708gb_Gaming8G == soup_kabum_Rx5708gb_Gaming8G.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx5708gb_Gaming8G_Promocao) 
def Rx5704gb_XFXRSXXX():
    while True:
        if status_kabum_Rx5704gb_XFXRSXXX != 'PROMOÇÃO':
            print('{}'.format(status_kabum_Rx5704gb_XFXRSXXX))


        else:
            kabum_Rx5704gb_XFXRSXXX_Promocao = api.update_status('''

PROMOÇÃO:

Rx570 4gb (RSXXX - XFX)

''' + oldPrice_kabum_Rx5704gb_XFXRSXXX + '''
''' + newPrice_kabum_Rx5704gb_XFXRSXXX + '''


Preço: R$''' + price_kabum_Rx5704gb_XFXRSXXX + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx5704gb_XFXRSXXX + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx5704gb_XFXRSXXX == soup_kabum_Rx5704gb_XFXRSXXX.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx5704gb_XFXRSXXX_Promocao) 

def Rx590_Gaming8G():
    while True:
        if status_kabum_Rx590_Gaming8G != 'PROMOÇÃO':
            print('{}'.format(status_kabum_Rx590_Gaming8G))


        else:
            kabum_Rx590_Gaming8G_Promocao = api.update_status('''

PROMOÇÃO:

Rx590 (Gaming8G - Gigabyte)

''' + oldPrice_kabum_Rx590_Gaming8G + '''
''' + newPrice_kabum_Rx590_Gaming8G + '''


Preço: R$''' + price_kabum_Rx590_Gaming8G + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx590_Gaming8G + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx590_Gaming8G == soup_kabum_Rx590_Gaming8G.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx590_Gaming8G_Promocao) 



def Rx5500XT4gb_GamingOC():
    while True:
        if status_kabum_Rx5500XT4gb_GamingOC != 'PROMOÇÃO':
            print('{}'.format(status_kabum_Rx5500XT4gb_GamingOC))


        else:
            kabum_Rx5500XT4gb_GamingOC_Promocao = api.update_status('''

PROMOÇÃO:

Rx5500XT 4gb (GamingOC - Gigabyte)

''' + oldPrice_kabum_Rx5500XT4gb_GamingOC + '''
''' + newPrice_kabum_Rx5500XT4gb_GamingOC + '''


Preço: R$''' + price_kabum_Rx5500XT4gb_GamingOC + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx5500XT4gb_GamingOC + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx5500XT4gb_GamingOC == soup_kabum_Rx5500XT4gb_GamingOC.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx5500XT4gb_GamingOC_Promocao) 
        
def Rx5500XT8gb_Pulse():
    while True:
        if status_kabum_Rx5500XT8gb_Pulse != 'PROMOÇÃO':
                print('{}'.format(status_kabum_Rx5500XT8gb_Pulse))


        else:
            kabum_Rx5500XT8gb_Pulse_Promocao = api.update_status('''

PROMOÇÃO:

Rx5500XT 8gb (Pulse - Sapphire)

''' + oldPrice_kabum_Rx5500XT8gb_Pulse + '''
''' + newPrice_kabum_Rx5500XT8gb_Pulse + '''


Preço: R$''' + price_kabum_Rx5500XT8gb_Pulse + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_Rx5500XT8gb_Pulse + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_Rx5500XT8gb_Pulse == soup_kabum_Rx5500XT8gb_Pulse.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_Rx5500XT8gb_Pulse_Promocao) 
        
def i39100F():
    while True:
        if status_kabum_i39100F != 'PROMOÇÃO':
                print('{}'.format(status_kabum_i39100F))


        else:
            kabum_i39100F_Promocao = api.update_status('''

PROMOÇÃO:

i3 9100F

''' + oldPrice_kabum_i39100F + '''
''' + newPrice_kabum_i39100F + '''


Preço: R$''' + price_kabum_i39100F + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i39100F + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i39100F == soup_kabum_i39100F.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i39100F_Promocao)

def i59400F():
    while True:
        if status_kabum_i59400F != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i59400F))


        else:
            kabum_i59400F_Promocao = api.update_status('''

PROMOÇÃO:

i5 9400F

''' + oldPrice_kabum_i59400F + '''
''' + newPrice_kabum_i59400F + '''


Preço: R$''' + price_kabum_i59400F + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i59400F + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i59400F == soup_kabum_i59400F.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i59400F_Promocao)


def i59600K():
    while True:
        if status_kabum_i59600K != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i59600K))


        else:
            kabum_i59600K_Promocao = api.update_status('''

PROMOÇÃO:

i5 9600K

''' + oldPrice_kabum_i59600K + '''
''' + newPrice_kabum_i59600K + '''


Preço: R$''' + price_kabum_i59600K + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i59600K + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i59600K == soup_kabum_i59600K.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i59600K_Promocao)
def i59600KF():
    while True:
        if status_kabum_i59600KF != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i59600KF))


        else:
            kabum_i59600KF_Promocao = api.update_status('''

PROMOÇÃO:

i5 9600KF

''' + oldPrice_kabum_i59600KF + '''
''' + newPrice_kabum_i59600KF + '''


Preço: R$''' + price_kabum_i59600KF + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i59600KF + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i59600KF == soup_kabum_i59600KF.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i59600KF_Promocao)

def i79700F():
    while True:
        if status_kabum_i79700F != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i79700F))


        else:
            kabum_i79700F_Promocao = api.update_status('''

PROMOÇÃO:

i7 9700F

''' + oldPrice_kabum_i79700F + '''
''' + newPrice_kabum_i79700F + '''


Preço: R$''' + price_kabum_i79700F + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i79700F + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i79700F == soup_kabum_i79700F.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i79700F_Promocao)


def i79700K():
    while True:
        if status_kabum_i79700K != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i79700K))


        else:
            kabum_i79700K_Promocao = api.update_status('''

PROMOÇÃO:

i7 9700K

''' + oldPrice_kabum_i79700K + '''
''' + newPrice_kabum_i79700K + '''


Preço: R$''' + price_kabum_i79700K + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i79700K + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i79700K == soup_kabum_i79700K.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i79700K_Promocao)

def i79700KF():
    while True:
        if status_kabum_i79700KF != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i79700KF))


        else:
            kabum_i79700KF_Promocao = api.update_status('''

PROMOÇÃO:

i7 9700KF

''' + oldPrice_kabum_i79700KF + '''
''' + newPrice_kabum_i79700KF + '''


Preço: R$''' + price_kabum_i79700KF + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i79700KF + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i79700KF == soup_kabum_i79700KF.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i79700KF_Promocao)

def i99900KF():
    while True:
        if status_kabum_i99900KF != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i99900KF))


        else:
            kabum_i99900KF_Promocao = api.update_status('''

PROMOÇÃO:

i9 9900KF

''' + oldPrice_kabum_i99900KF + '''
''' + newPrice_kabum_i99900KF + '''


Preço: R$''' + price_kabum_i99900KF + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i99900KF + '''




                                                                    ''')
            
            break
            while True:
                if status_kabum_i99900KF == soup_kabum_i99900KF.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i99900KF_Promocao)
          
                
                    
def i99900K():
    while True:
        if status_kabum_i99900K != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i99900K))


        else:
            kabum_i99900K_Promocao = api.update_status('''

PROMOÇÃO:

i9 9900K

''' + oldPrice_kabum_i99900K + '''
''' + newPrice_kabum_i99900K + '''


Preço: R$''' + price_kabum_i99900K + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i99900K + '''



                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i99900K == soup_kabum_i99900K.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i99900K_Promocao)


def i310100():
    while True:
        if status_kabum_i310100 != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i310100))


        else:
            kabum_i310100_Promocao = api.update_status('''

PROMOÇÃO:

i3 10100

''' + oldPrice_kabum_i310100 + '''
''' + newPrice_kabum_i310100 + '''


Preço: R$''' + price_kabum_i310100 + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i310100 + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i310100 == soup_kabum_i310100.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i310100_Promocao)

def i510400():
    while True:
        if status_kabum_i510400 != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i510400))


        else:
            kabum_i510400_Promocao = api.update_status('''

PROMOÇÃO:

i5 10400

''' + oldPrice_kabum_i510400 + '''
''' + newPrice_kabum_i510400 + '''


Preço: R$''' + price_kabum_i510400 + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i510400 + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i510400 == soup_kabum_i510400.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i510400_Promocao)


def i510400F():
    while True:
        if status_kabum_i510400F != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i510400F))


        else:
            kabum_i510400F_Promocao = api.update_status('''

PROMOÇÃO:

i5 10400F

''' + oldPrice_kabum_i510400F + '''
''' + newPrice_kabum_i510400F + '''


Preço: R$''' + price_kabum_i510400F + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i510400F + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i510400F == soup_kabum_i510400F.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i510400F_Promocao)


def i710700():
    while True:
        if status_kabum_i710700 != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i710700))


        else:
            kabum_i710700_Promocao = api.update_status('''

PROMOÇÃO:

i7 10700

''' + oldPrice_kabum_i710700 + '''
''' + newPrice_kabum_i710700 + '''


Preço: R$''' + price_kabum_i710700 + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i710700 + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i710700 == soup_kabum_i710700.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i710700_Promocao)
def i710700K():
    while True:
        if status_kabum_i710700K != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i710700K))


        else:
            kabum_i710700K_Promocao = api.update_status('''

PROMOÇÃO:

i7 10700K

''' + oldPrice_kabum_i710700K + '''
''' + newPrice_kabum_i710700K + '''


Preço: R$''' + price_kabum_i710700K + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i710700K + '''



                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i710700K == soup_kabum_i710700K.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i710700K_Promocao)



def i910900K():
    while True:
        if status_kabum_i910900K != 'PROMOÇÃO':
            print('{}'.format(status_kabum_i910900K))


        else:
            kabum_i910900K_Promocao = api.update_status('''

PROMOÇÃO:

i9 10900K

''' + oldPrice_kabum_i910900K + '''
''' + newPrice_kabum_i910900K + '''


Preço: R$''' + price_kabum_i910900K + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_i910900K + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_i910900K == soup_kabum_i910900K.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_i910900K_Promocao)




def GTX1650Super_XCBlackGaming():
    while True:
        if status_kabum_GTX1650Super_XCBlackGaming != 'PROMOÇÃO':
            print('{}'.format(status_kabum_GTX1650Super_XCBlackGaming))


        else:
            kabum_GTX1650Super_XCBlackGaming_Promocao = api.update_status('''

PROMOÇÃO:

GTX 1650Super (XC BlackGaming - EVGA)

''' + oldPrice_kabum_GTX1650Super_XCBlackGaming + '''
''' + newPrice_kabum_GTX1650Super_XCBlackGaming + '''


Preço: R$''' + price_kabum_GTX1650Super_XCBlackGaming + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_GTX1650Super_XCBlackGaming + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_GTX1650Super_XCBlackGaming == soup_kabum_GTX1660_GigabyteOC.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_GTX1650Super_XCBlackGaming_Promocao)

def GTX1660_GigabyteOC():
    while True:
        if status_kabum_GTX1660_GigabyteOC != 'PROMOÇÃO':
            print('{}'.format(status_kabum_GTX1660_GigabyteOC))


        else:
            kabum_GTX1660_GigabyteOC_Promocao = api.update_status('''

PROMOÇÃO:

GTX 1660 (GigabyteOC - Gigabyte)

''' + oldPrice_kabum_GTX1660_GigabyteOC + '''
''' + newPrice_kabum_GTX1660_GigabyteOC + '''


Preço: R$''' + price_kabum_GTX1660_GigabyteOC + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_GTX1660_GigabyteOC + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_GTX1660_GigabyteOC == soup_kabum_GTX1660Super_GigabyteOC.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_GTX1660_GigabyteOC_Promocao)


def GTX1660Super_GigabyteOC():
    while True:
        if status_kabum_GTX1660Super_GigabyteOC != 'PROMOÇÃO':
            print('{}'.format(status_kabum_GTX1660Super_GigabyteOC))


        else:
            kabum_GTX1660Super_GigabyteOC_Promocao = api.update_status('''

PROMOÇÃO:

GTX 1660Super (GigabyteOC - Gigabyte)

''' + oldPrice_kabum_GTX1660Super_GigabyteOC + '''
''' + newPrice_kabum_GTX1660Super_GigabyteOC + '''


Preço: R$''' + price_kabum_GTX1660Super_GigabyteOC + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_GTX1660Super_GigabyteOC + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_GTX1660Super_GigabyteOC == soup_kabum_GTX1660Super_GigabyteOC.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_GTX1660Super_GigabyteOC_Promocao)



def GTX1660Ti_GamingX():
    while True:
        if status_kabum_GTX1660Ti_GamingX != 'PROMOÇÃO':
            print('{}'.format(status_kabum_GTX1660Ti_GamingX))


        else:
            kabum_GTX1660Ti_GamingX_Promocao = api.update_status('''

PROMOÇÃO:

GTX 1660Super (GigabyteOC - Gigabyte)

''' + oldPrice_kabum_GTX1660Ti_GamingX + '''
''' + newPrice_kabum_GTX1660Ti_GamingX + '''


Preço: R$''' + price_kabum_GTX1660Ti_GamingX + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_GTX1660Ti_GamingX + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_GTX1660Ti_GamingX == soup_kabum_GTX1660Ti_GamingX.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_GTX1660Ti_GamingX_Promocao)

def RTX2060_Gaming():
        while True:
            if status_kabum_RTX2060_Gaming != 'PROMOÇÃO':
                print('{}'.format(status_kabum_RTX2060_Gaming))


            else:
                kabum_RTX2060_Gaming_Promocao = api.update_status('''

PROMOÇÃO:

RTX2060 (Gaming - MSI)

''' + oldPrice_kabum_RTX2060_Gaming + '''
''' + newPrice_kabum_RTX2060_Gaming + '''


Preço: R$''' + price_kabum_RTX2060_Gaming + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_RTX2060_Gaming + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_RTX2060_Gaming == soup_kabum_RTX2060_Gaming.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_RTX2060_Gaming_Promocao)

def RTX2060_GamingOCPro():
    while True:
        if status_kabum_RTX2060_GamingOCPro != 'PROMOÇÃO':
            print('{}'.format(status_kabum_RTX2060_GamingOCPro))


        else:
            kabum_RTX2060_GamingOCPro_Promocao = api.update_status('''

PROMOÇÃO:

RTX2060 (GamingOCPro - Gigabyte)

''' + oldPrice_kabum_RTX2060_GamingOCPro + '''
''' + newPrice_kabum_RTX2060_GamingOCPro + '''


Preço: R$''' + price_kabum_RTX2060_GamingOCPro + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_RTX2060_GamingOCPro + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabum_RTX2060_GamingOCPro == soup_kabum_RTX2060_GamingOCPro.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_RTX2060_GamingOCPro_Promocao)

def RTX2060Super_ROGStrix():
    while True:
        if status_kabum_RTX2060Super_ROGStrix != 'PROMOÇÃO':
            print('{}'.format(status_kabum_RTX2060Super_ROGStrix))


        else:
            kabum_RTX2060Super_ROGStrix_Promocao = api.update_status('''
PROMOÇÃO:

RTX 2060Super (ROGStrix - Asus)

''' + oldPrice_kabum_RTX2060Super_ROGStrix + '''
''' + newPrice_kabum_RTX2060Super_ROGStrix + '''


Preço: R$''' + price_kabum_RTX2060Super_ROGStrix + '''
Á vista no boleto bancário com 15% de desconto



''' + kabum_RTX2060Super_ROGStrix + '''


                                                                    ''')
            
            break
            while True:
          
                if status_kabumRTX2060Super_ROGStrix == soup_kabum_RTX2060Super_ROGStrix.find(src="https://static.kabum.com.br/conteudo/temas/001/imagens/descricao/bot_disponibilidade_on.gif").get("alt"):
                    api.destroy_status(kabum_RTX2060Super_ROGStrix_Promocao)


        
if __name__ == '__main__':
    Ryzen51600AF = Thread(target = R51600AF)
    Ryzen32200G = Thread(target = R32200G)
    Ryzen52600 = Thread(target = R52600)
    Ryzen53600 = Thread(target = R53600)
    Ryzen72700 = Thread(target = R72700)
    Ryzen73700X = Thread(target = R73700X)
    Ryzen73800X = Thread(target = R73800X)
    Ryzen93900X = Thread(target = R93900X)
    Ryzen33300X = Thread(target = R33300X)
    Ryzen33200G = Thread(target = R33200G)
    Ryzen53600X = Thread(target = R53600X)
    RadeonRx5704gb_PhantomGaming = Thread(target = Rx5704gb_PhantomGaming)
    RadeonRx5708gb_PhantomGaming = Thread(target = Rx5708gb_PhantomGaming)
    RadeonRx5704gb_Gaming4G = Thread(target = Rx5704gb_Gaming4G)
    RadeonRx5708gb_Gaming8G = Thread(target = Rx5708gb_Gaming8G)
    RadeonRx5704gb_XFXRSXXX = Thread(target = Rx5704gb_XFXRSXXX)
    RadeonRx590_Gaming8G = Thread(target = Rx590_Gaming8G)
    RadeonRx5500XT4gb_GamingOC = Thread(target = Rx5500XT4gb_GamingOC)
    RadeonRx5500XT8gb_Pulse = Thread(target = Rx5500XT8gb_Pulse)
    

    Ryzen51600AF.start()
    Ryzen32200G.start()
    Ryzen52600.start()
    Ryzen53600.start()
    Ryzen72700.start()
    Ryzen73700X.start()
    Ryzen73800X.start()
    Ryzen93900X.start()
    Ryzen33300X.start()
    Ryzen33200G.start()
    Ryzen53600X.start()
    RadeonRx5704gb_PhantomGaming.start()
    RadeonRx5708gb_PhantomGaming.start()
    RadeonRx5704gb_Gaming4G.start()
    RadeonRx5708gb_Gaming8G.start()
    RadeonRx5704gb_XFXRSXXX.start()
    RadeonRx590_Gaming8G.start()
    RadeonRx5500XT4gb_GamingOC.start()
    RadeonRx5500XT8gb_Pulse.start()


    Corei39100F = Thread(target = i39100F)
    Corei99900KF = Thread(target = i99900KF)
    Corei59400F = Thread(target = i59400F)
    Corei59600K = Thread(target = i59600K)
    Corei59600KF = Thread(target = i59600KF)
    Corei79700F = Thread(target = i79700F)
    Corei79700K = Thread(target = i79700K)
    Corei79700KF = Thread(target = i79700KF)
    Corei99900KF = Thread(target = i99900KF)
    Corei99900K = Thread(target = i99900K)
    Corei310100 = Thread(target = i310100)
    Corei510400 = Thread(target = i310100)
    Corei510400F = Thread(target = i510400F)
    Corei710700 = Thread(target = i710700)
    Corei710700K = Thread(target = i710700K)
    Corei910900K = Thread(target = i910900K)


    Corei39100F.start()
    Corei99900KF.start()
    Corei59400F.start()
    Corei59600K.start()
    Corei59600KF.start()
    Corei79700F.start()
    Corei79700K.start()
    Corei79700KF.start()
    Corei99900KF.start()
    Corei99900K.start()
    Corei310100.start()
    Corei510400.start()
    Corei510400F.start()
    Corei710700.start()
    Corei710700K.start()
    Corei910900K.start()


    GeForceGTX1650Super_XCBlackGaming = Thread(target = GTX1650Super_XCBlackGaming)
    GeForceGTX1660_GigabyteOC = Thread(target = GTX1660_GigabyteOC)
    GeForceGTX1660Super_GigabyteOC = Thread(target = GTX1660Super_GigabyteOC)
    GeForceGTX1660Ti_GamingX = Thread(target = GTX1660Ti_GamingX)
    GeForceRTX2060_Gaming = Thread(target = RTX2060_Gaming)
    GeForceRTX2060_GamingOCPro = Thread(target = RTX2060_GamingOCPro)
    GeForceRTX2060Super_ROGStrix = Thread(target = RTX2060Super_ROGStrix)
    




    GeForceGTX1650Super_XCBlackGaming.start()
    GeForceGTX1660_GigabyteOC.start()
    GeForceGTX1660Super_GigabyteOC.start()
    GeForceGTX1660Ti_GamingX.start()
    GeForceRTX2060_Gaming.start()
    GeForceRTX2060_GamingOCPro.start()
    GeForceRTX2060Super_ROGStrix.start()























